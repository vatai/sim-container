#!/usr/bin/env python

from pathlib import Path
from pprint import pprint
from typing import Optional

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import Token

# from openpyxl.workbook.workbook import Workbook
# from openpyxl.worksheet.worksheet import Worksheet

LINES = []
OUTPUT_DICT = {}
PROCESSED_CELLS = set()
WORKSHEET_STACK = []

FAPP_XML_OBJ = "fapp_xml"
SPECIAL_FAPP_XML_CALL = {
    "C4": f"{FAPP_XML_OBJ}.get_counter_timer_freq()",
    "G4": f"{FAPP_XML_OBJ}.get_measured_time()",
    "G5": f"{FAPP_XML_OBJ}.get_node_name()",
    "G10": f"{FAPP_XML_OBJ}.get_process_no()",
    "G11": f"{FAPP_XML_OBJ}.get_cmg_no()",
    "G12": f"{FAPP_XML_OBJ}.get_measured_region()",
    "G14": f"{FAPP_XML_OBJ}.get_vector_length()",
}
HEADER_ROW = 29
TOP_ROW = 30
BOTOM_ROW = 41
LEFT_COLUMN = 29
RIGHT_COLUMN = 334

INFIX_OP_MAP = {
    "=": "==",
    "^": "**",
}

DEBUG = True
DEBUG = False


def event_cell(cell_id):
    dbg_print(f"== BEGIN EVENT_CELL({cell_id})")
    cell = get_cell(cell_id)
    if isinstance(cell, tuple):
        results = [event_cell(cell_to_id(c)) for c in cell]
        result = any(results)
        assert result == all(results)
    else:
        col = cell.col_idx
        row = cell.row
        result = (
            LEFT_COLUMN <= col
            and col <= RIGHT_COLUMN
            and TOP_ROW <= row
            and row <= BOTOM_ROW
        )
    dbg_print(f"== END EVENT_CELL -> {result}")
    return result


def get_coords_on_right(cell_loc: str):
    worksheet = WORKBOOK["report"]
    cell = worksheet[cell_loc]
    row, col = cell.row, cell.col_idx

    while not isinstance(worksheet.cell(row, col), MergedCell):
        col += 1
    col += 1
    return worksheet.cell(row, col).coordinate


# TRASH #


def full_cell_id(cell_id: str) -> str:
    prefix = WORKSHEET_STACK[-1] if WORKSHEET_STACK else "report"
    if "!" not in cell_id:
        return f"{prefix}!{cell_id}"
    return cell_id


def cell_to_id(cell: Cell) -> str:
    return cell[0].parent.title + "!" + cell[0].coordinate


def cell_id_to_var(cell_id: str) -> str:
    dbg_print(f"BEGIN CELL_ID_TO_VAR({cell_id})")
    cells = get_cell(cell_id)
    if isinstance(cells, tuple):
        result = ", ".join([cell_id_to_var(cell_to_id(cell)) for cell in cells])
        result = f"[{result}]"
    else:
        result = cell_id.replace("!", "_").replace("$", "")
    dbg_print(f"END CELL_ID_TO_VAR -> {result}")
    return result


def get_cell(cell_id: str) -> Cell:
    dbg_print(f"BEGIN GET_CELL({cell_id})")
    cell_id = full_cell_id(cell_id)
    # if "!" not in cell_id:
    #     cell_id = WORKSHEET_STACK[-1] + "!" + cell_id
    ws, cell = cell_id.split("!")
    cell = WORKBOOK[ws][cell]
    if isinstance(cell, MergedCell):
        print("WARNING: MergedCell!")
    dbg_print(f"END GET_CELL -> {cell}")
    return cell


def dbg_print(msg: str) -> None:
    if DEBUG:
        print(msg)


def unknown_type(token: Token) -> None:
    msg = f"ERROR: Unknown type {token.type}"
    raise Exception(msg)


def unknown_subtype(token: Token) -> None:
    msg = f"ERROR: Unknown subtype {token.subtype} (of token type {token.type})"
    raise Exception(msg)


def unknown_func(token: Token) -> None:
    msg = f"ERROR: Unknown FUNC {token.value}"
    raise Exception(msg)


def assert_sep_comma(token: Token) -> None:
    assert token.type == Token.SEP and token.value == ","


def assert_func_close(token: Token) -> None:
    assert token.type == Token.FUNC and token.subtype == Token.CLOSE


def get_raw(cell_id: str) -> Optional[str]:
    dbg_print(f"BEGIN GET_RAW({cell_id})")
    result = None
    cell_id = cell_id.replace("$", "")
    if ":" in cell_id:
        cells = get_cell(cell_id)
        results = [get_raw(cell_to_id(cell)) for cell in cells]
        if any(results):
            result = f"[{', '.join(results)}]"
    else:
        cell_id = full_cell_id(cell_id)
        # if "!" not in cell_id:
        #     cell_id = WORKSHEET_STACK[-1] + "!" + cell_id
        ws, cell = cell_id.split("!")
        if ws == "label":
            value = WORKBOOK[ws][cell].value
            result = f"'{value}'"
        elif ws == "data":
            if cell in SPECIAL_FAPP_XML_CALL:
                result = SPECIAL_FAPP_XML_CALL[cell]
            elif event_cell(cell_id):
                cell_obj = WORKBOOK[ws][cell]
                col = cell_obj.col_idx - 1
                event_name = WORKBOOK[ws][HEADER_ROW][col].value
                thread_id = cell_obj.row - HEADER_ROW - 1
                result = f"{FAPP_XML_OBJ}.get_event('{event_name}', {thread_id})"
    dbg_print(f"END GET_RAW -> {result}")
    return result


def parse_tokens(tokens: list[Token], cur: int) -> (str, int):
    dbg_print(f"BEGIN PARSE_TOKENS({tokens}, {cur}")
    result = ""
    while cur < len(tokens):
        token = tokens[cur]
        if token.type == Token.OPERAND:
            if token.subtype == Token.RANGE:
                cell_id = full_cell_id(token.value)
                raw = get_raw(cell_id)
                if raw:
                    result += raw
                else:
                    cell_to_inst(cell_id)
                    result += cell_id_to_var(cell_id)
            elif token.subtype == Token.TEXT:
                result += token.value
            elif token.subtype == Token.NUMBER:
                result += token.value
            else:
                unknown_subtype(token)
        elif token.type == Token.FUNC:
            if token.subtype == Token.OPEN:
                if token.value == "IF(":
                    cond, cur = parse_tokens(tokens, cur + 1)
                    assert_sep_comma(tokens[cur])
                    true_val, cur = parse_tokens(tokens, cur + 1)
                    assert_sep_comma(tokens[cur])
                    false_val, cur = parse_tokens(tokens, cur + 1)
                    assert_func_close(tokens[cur])
                    result += f"({true_val}) if ({cond}) else ({false_val})"
                elif token.value == "OR(":
                    ops, cur = parse_tokens(tokens, cur + 1)
                    while tokens[cur].type == Token.SEP and tokens[cur].value == ",":
                        tmp, cur = parse_tokens(tokens, cur + 1)
                        ops += f", {tmp}"
                    assert_func_close(tokens[cur])
                    result += f"(any([{ops}]))"
                elif token.value == "COUNT(":
                    cells, cur = parse_tokens(tokens, cur + 1)
                    assert_func_close(tokens[cur])
                    result += f"(sum(1 for e in {cells} if e !=''))"
                else:
                    unknown_func(token)
            elif token.subtype == Token.CLOSE:
                break
            else:
                unknown_subtype(token)
        elif token.type == Token.OP_IN:
            op_name = token.value
            if token.value in INFIX_OP_MAP:
                op_name = INFIX_OP_MAP[token.value]
            result += op_name
        elif token.type == Token.SEP:
            break
        else:
            unknown_type(token)
        cur += 1
    dbg_print(f"END PARSE_TOKENS - > {result}, {cur}")
    return result, cur


def parse_cell(cell: Cell) -> str:
    dbg_print(f"BEGIN PARSE_CELL({cell} with value {cell.value})")
    WORKSHEET_STACK.append(cell.parent.title)
    tokens = Tokenizer(cell.value).items
    value, _ = parse_tokens(tokens, 0)
    WORKSHEET_STACK.pop()
    dbg_print(f"END PARSE_CELL -> {value}")
    return value


def cell_to_inst(cell_id: str) -> None:
    dbg_print(f"BEGIN CELL_TO_INST({cell_id})")
    cell_id = full_cell_id(cell_id)
    cell = get_cell(cell_id)
    if isinstance(cell, tuple):
        for c in cell:
            cell_to_inst(cell_to_id(c))
    else:
        cell_var = cell_id_to_var(cell_id)
        if cell_var in PROCESSED_CELLS:
            return
        cell_val = parse_cell(cell)
        line = f"{cell_var} = {cell_val}"
        dbg_print(f"END CELL_TO_INST: LINES.append({line})")
        PROCESSED_CELLS.add(cell_var)
        LINES.append(line)


# main prelude
if "WORKBOOK" not in locals():
    filename = Path(
        "~/Sync/tmp/work/fapp-xmls/gemver_LARGE.fapp.report/cpu_pa_report.xlsm"
    ).expanduser()
    WORKBOOK = openpyxl.load_workbook(filename)


def add_key_single_value_pair(key: str, value: str) -> None:
    key = full_cell_id(key)
    value = full_cell_id(value)
    cell_to_inst(key)
    cell_to_inst(value)
    OUTPUT_DICT[cell_id_to_var(key)] = cell_id_to_var(value)


def create_program() -> str:
    with open("fapp_top.py.in") as top:
        result = top.readlines()
    result += LINES

    line = "result={"
    for key, value in OUTPUT_DICT.items():
        line += f"{key}: {value}, "
    line += "}"
    result.append(line)

    with open("fapp_bottom.py.in") as top:
        result += top.readlines()
    return "\n".join(result)


def main():
    add_key_single_value_pair("A3", "C3")
    add_key_single_value_pair("A4", "C4")
    add_key_single_value_pair("H3", "J3")
    add_key_single_value_pair("H4", "J4")
    add_key_single_value_pair("H5", "J5")
    add_key_single_value_pair("O3", "Q3")
    add_key_single_value_pair("O4", "Q4")

    assert WORKSHEET_STACK == []

    program = create_program()

    with open("read_fapp_xmls.out.py", "w") as out:
        out.write(program)

    # print(LINES)
    print(program)
    exec(program)


main()
